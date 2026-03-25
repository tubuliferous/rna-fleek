#!/usr/bin/env python3
"""
Download and preprocess CellMarker2 database into a compact JSON lookup.

Creates cell_markers.json with structure:
{
  "Human": {
    "tissue_name": {
      "Cell Type Name": ["GENE1", "GENE2", ...],
      ...
    }
  },
  "Mouse": { ... },
  "_global": {
    "Human": {
      "Cell Type Name": ["GENE1", "GENE2", ...],  # union across all tissues
      ...
    }
  }
}

Usage:
    python download_markers.py
    python download_markers.py --output my_markers.json

Requirements:
    pip install requests openpyxl
"""

import argparse
import json
import os
import sys
import tempfile
from collections import defaultdict
from pathlib import Path


def download_cellmarker2(output_path="cell_markers.json"):
    """Download CellMarker2 database and convert to JSON."""
    try:
        import requests
    except ImportError:
        print("Error: 'requests' package required. Install with: pip install requests")
        sys.exit(1)
    try:
        import openpyxl
    except ImportError:
        print("Error: 'openpyxl' package required. Install with: pip install openpyxl")
        sys.exit(1)

    # CellMarker 2.0 download URL
    url = "http://bio-bigdata.hrbmu.edu.cn/CellMarker/CellMarker_download_files/file/Cell_marker_All.xlsx"
    
    print(f"Downloading CellMarker2 database...")
    print(f"  URL: {url}")
    
    # Download to temp file
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        r = requests.get(url, timeout=120, stream=True)
        r.raise_for_status()
        total = int(r.headers.get("content-length", 0))
        downloaded = 0
        for chunk in r.iter_content(chunk_size=65536):
            tmp.write(chunk)
            downloaded += len(chunk)
            if total:
                print(f"\r  Downloaded {downloaded/(1024*1024):.1f} / {total/(1024*1024):.1f} MB", end="", flush=True)
        print()
        tmp.close()
        
        print("  Parsing Excel file...")
        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
        ws = wb.active
        
        # Find column indices from header row
        headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        # Map expected columns
        col_map = {}
        for i, h in enumerate(headers):
            if "species" in h:
                col_map["species"] = i
            elif "tissue" in h and "type" in h:
                col_map["tissue"] = i
            elif "cell_name" in h or "cellname" in h or h == "cell name":
                col_map["cell_name"] = i
            elif h in ("symbol", "gene_symbol", "marker"):
                col_map["symbol"] = i
            elif "cell_type" in h or "celltype" in h:
                col_map["cell_type"] = i
        
        print(f"  Found columns: {col_map}")
        
        if "species" not in col_map or "cell_name" not in col_map or "symbol" not in col_map:
            # Try alternative: just use positional
            print("  Warning: Could not find all columns by name, trying positional...")
            # Common layout: species(0), tissue(1), ..., cell_name(5), ..., Symbol(8)
            # Let's just print first row to debug
            for row in ws.iter_rows(min_row=1, max_row=2):
                print("  ", [str(c.value)[:30] for c in row])
            print("  Please check the file format and update column mapping.")
            wb.close()
            return
        
        # Parse rows
        by_species_tissue = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
        by_species_global = defaultdict(lambda: defaultdict(set))
        
        n_rows = 0
        for row in ws.iter_rows(min_row=2):
            vals = [str(cell.value or "").strip() for cell in row]
            if len(vals) <= max(col_map.values()):
                continue
            
            species = vals[col_map["species"]]
            tissue = vals[col_map.get("tissue", col_map.get("cell_type", 0))]
            cell_name = vals[col_map["cell_name"]]
            symbol = vals[col_map["symbol"]].upper()
            
            if not species or not cell_name or not symbol or symbol == "NONE":
                continue
            
            # Clean up
            species = species.strip().capitalize()
            if species not in ("Human", "Mouse"):
                continue
            
            cell_name = cell_name.strip()
            tissue = tissue.strip() if tissue else "Unknown"
            
            # Some entries have multiple genes separated by commas or semicolons
            for gene in symbol.replace(";", ",").split(","):
                gene = gene.strip().upper()
                if gene and len(gene) > 1:
                    by_species_tissue[species][tissue][cell_name].add(gene)
                    by_species_global[species][cell_name].add(gene)
                    n_rows += 1
        
        wb.close()
        print(f"  Parsed {n_rows:,} marker entries")
        
        # Convert sets to sorted lists for JSON
        result = {}
        for species in by_species_tissue:
            result[species] = {}
            for tissue in by_species_tissue[species]:
                result[species][tissue] = {}
                for cell_name, genes in by_species_tissue[species][tissue].items():
                    result[species][tissue][cell_name] = sorted(genes)
        
        result["_global"] = {}
        for species in by_species_global:
            result["_global"][species] = {}
            for cell_name, genes in by_species_global[species].items():
                result["_global"][species][cell_name] = sorted(genes)
        
        # Stats
        for species in ["Human", "Mouse"]:
            if species in result.get("_global", {}):
                n_types = len(result["_global"][species])
                n_genes = sum(len(g) for g in result["_global"][species].values())
                print(f"  {species}: {n_types:,} cell types, {n_genes:,} total marker genes")
        
        # Save
        with open(output_path, "w") as f:
            json.dump(result, f, separators=(",", ":"))
        
        size_mb = os.path.getsize(output_path) / 1e6
        print(f"  Saved to {output_path} ({size_mb:.1f} MB)")
        
    finally:
        os.unlink(tmp.name)


def main():
    parser = argparse.ArgumentParser(description="Download CellMarker2 database")
    parser.add_argument("--output", type=str, default="cell_markers.json",
                        help="Output JSON file path")
    args = parser.parse_args()
    download_cellmarker2(args.output)


if __name__ == "__main__":
    main()